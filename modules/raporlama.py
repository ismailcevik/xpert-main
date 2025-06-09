import os
import json
import pandas as pd
from flask import (
    Blueprint, request, render_template, send_file, current_app as app
)
from datetime import datetime
import plotly.graph_objs as go
from io import BytesIO
from openpyxl import load_workbook

bp = Blueprint('raporlama', __name__, url_prefix='/raporlama')

DATA_FILE = 'database.xlsx'  # Örnek veri dosyası
CRITERIA_DIR = 'criteria_jsons'
TEMPLATE_DIR = 'excel_templates'

def load_data():
    df = pd.read_excel(DATA_FILE)
    df['RealDate'] = pd.to_datetime(df['RealDate'])
    return df

def filter_dataframe(df, start_dt, end_dt, filters):
    # Tarih aralığı filtresi
    if start_dt:
        df = df[df['RealDate'] >= start_dt]
    if end_dt:
        df = df[df['RealDate'] <= end_dt]
    
    # Kolon bazlı filtreler (örn: >100, <=200, =150, !=50, regex destekli değil)
    for col, expr in filters.items():
        expr = expr.strip()
        if expr:
            import re
            m = re.match(r'(<=|>=|<|>|=|!=)\s*(.+)', expr)
            if m:
                op, val = m.group(1), m.group(2)
                try:
                    val = float(val)
                except:
                    val = val.strip()
                if op == '>':
                    df = df[df[col] > val]
                elif op == '>=':
                    df = df[df[col] >= val]
                elif op == '<':
                    df = df[df[col] < val]
                elif op == '<=':
                    df = df[df[col] <= val]
                elif op == '=':
                    df = df[df[col] == val]
                elif op == '!=':
                    df = df[df[col] != val]
            else:
                try:
                    val = float(expr)
                    df = df[df[col] == val]
                except:
                    pass
    return df

def get_hourly_avg(df, columns):
    df['Hour'] = df['RealDate'].dt.floor('H')
    grouped = df.groupby('Hour')[columns].mean().reset_index()
    grouped['Hour'] = grouped['Hour'].dt.strftime('%Y-%m-%d %H:%M')
    return grouped

def create_trend_scatter(df, columns):
    fig = go.Figure()
    for col in columns:
        fig.add_trace(go.Scatter(
            x=df['RealDate'], y=df[col],
            mode='lines+markers',
            name=col,
            hovertemplate='%{y}<extra></extra>'
        ))
    fig.update_layout(
        margin=dict(l=20, r=20, t=30, b=20),
        hovermode='x unified',
        xaxis_title='Tarih',
        yaxis_title='Değer',
        height=350
    )
    return fig.to_html(full_html=False)

def create_hourly_bar_charts(hourly_df, columns):
    charts_html = ''
    for col in columns:
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=hourly_df['Hour'],
            y=hourly_df[col],
            name=col
        ))
        fig.update_layout(
            title=f'Saatlik Ortalama - {col}',
            margin=dict(l=20, r=20, t=40, b=30),
            height=300,
            xaxis_title='Saat',
            yaxis_title=col
        )
        charts_html += fig.to_html(full_html=False)
    return charts_html

def save_criteria(name, criteria):
    os.makedirs(CRITERIA_DIR, exist_ok=True)
    file_path = os.path.join(CRITERIA_DIR, f'{name}.json')
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(criteria, f, ensure_ascii=False, indent=2)

def load_criteria(name):
    file_path = os.path.join(CRITERIA_DIR, f'{name}.json')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None

@bp.route('/', methods=['GET', 'POST'])
def index():
    df = load_data()
    columns = [col for col in df.columns if col != 'RealDate']

    criteria_list = [f[:-5] for f in os.listdir(CRITERIA_DIR) if f.endswith('.json')] if os.path.exists(CRITERIA_DIR) else []

    # Varsayılan değerler
    start_dt = df['RealDate'].min()
    end_dt = df['RealDate'].max()
    filters = {col: '' for col in columns}
    selected_columns = columns.copy()
    load_name = None
    data = None
    trend_html = None
    hourly_avg_data = None
    hourly_charts_html = None

    if request.method == 'POST':
        # Kriter yükleme işlemi
        load_name = request.form.get('load_criteria')
        if load_name and load_name.strip():
            crit = load_criteria(load_name)
            if crit:
                # Kriterden yüklenen değerler
                try:
                    if crit.get('start_datetime'):
                        start_dt = datetime.fromisoformat(crit.get('start_datetime'))
                except:
                    pass
                try:
                    if crit.get('end_datetime'):
                        end_dt = datetime.fromisoformat(crit.get('end_datetime'))
                except:
                    pass
                
                # Filtreleri yükle
                saved_filters = crit.get('filters', {})
                filters = {col: saved_filters.get(col, '') for col in columns}
                
                # Sütun seçimlerini yükle
                saved_columns = crit.get('columns', [])
                if saved_columns:
                    selected_columns = [col for col in saved_columns if col in columns]
                else:
                    selected_columns = columns.copy()

        # Kriter kaydetme işlemi
        if request.form.get('save_criteria') and request.form.get('criteria_name'):
            name = request.form.get('criteria_name').strip()
            if name:
                # Önce manuel form verilerini al
                start_str = request.form.get('start_datetime')
                end_str = request.form.get('end_datetime')
                try:
                    if start_str:
                        start_dt = datetime.fromisoformat(start_str)
                    if end_str:
                        end_dt = datetime.fromisoformat(end_str)
                except Exception:
                    pass

                filters = {col: request.form.get(f'filter_{col}', '') for col in columns}
                selected_columns = request.form.getlist('columns') or columns
                
                # Kriterleri kaydet
                crit = {
                    'start_datetime': start_str,
                    'end_datetime': end_str,
                    'filters': filters,
                    'columns': selected_columns
                }
                save_criteria(name, crit)
                criteria_list = [f[:-5] for f in os.listdir(CRITERIA_DIR) if f.endswith('.json')]
                load_name = name  # Kaydedilen kriteri aktif göster

        # Manuel form verisi (sadece kriter yükleme veya kaydetme YOKSA)
        elif not (load_name and load_name.strip()):
            start_str = request.form.get('start_datetime')
            end_str = request.form.get('end_datetime')
            try:
                if start_str:
                    start_dt = datetime.fromisoformat(start_str)
                if end_str:
                    end_dt = datetime.fromisoformat(end_str)
            except Exception:
                pass

            filters = {col: request.form.get(f'filter_{col}', '') for col in columns}
            selected_columns = request.form.getlist('columns') or columns

        # Veri filtreleme ve görselleştirme (her durumda çalışır)
        if selected_columns:  # Sadece sütun seçimi varsa işle
            filtered_df = filter_dataframe(df, start_dt, end_dt, filters)
            data_df = filtered_df[['RealDate'] + selected_columns]
            data = data_df.to_dict(orient='records')
            trend_html = create_trend_scatter(data_df, selected_columns)
            hourly_df = get_hourly_avg(data_df, selected_columns)
            hourly_avg_data = hourly_df.to_dict(orient='records')
            hourly_charts_html = create_hourly_bar_charts(hourly_df, selected_columns)

    return render_template(
        'raporlama.html',
        columns=columns,
        selected_columns=selected_columns,
        filters=filters,
        start_dt_str=start_dt.strftime('%Y-%m-%dT%H:%M'),
        end_dt_str=end_dt.strftime('%Y-%m-%dT%H:%M'),
        data=data,
        trend_html=trend_html,
        hourly_avg_data=hourly_avg_data,
        hourly_charts_html=hourly_charts_html,
        criteria_list=criteria_list,
        load_name=load_name
    )

@bp.route('/export', methods=['POST'])
def export_excel():
    df = load_data()
    columns = request.form.getlist('columns')
    start_str = request.form.get('start_datetime')
    end_str = request.form.get('end_datetime')
    template_name = request.form.get('template_name', 'template.xlsx')

    try:
        start_dt = datetime.fromisoformat(start_str) if start_str else None
        end_dt = datetime.fromisoformat(end_str) if end_str else None
    except Exception:
        start_dt, end_dt = None, None

    filtered_df = df.copy()
    if start_dt:
        filtered_df = filtered_df[filtered_df['RealDate'] >= start_dt]
    if end_dt:
        filtered_df = filtered_df[filtered_df['RealDate'] <= end_dt]

    filtered_df = filtered_df[['RealDate'] + columns]

    template_path = os.path.join(TEMPLATE_DIR, template_name)
    if not os.path.isfile(template_path):
        bio = BytesIO()
        filtered_df.to_excel(bio, index=False)
        bio.seek(0)
        filename = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    wb = load_workbook(template_path)
    ws = wb.active

    for r_idx, row in enumerate(filtered_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')